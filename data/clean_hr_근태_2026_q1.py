# -*- coding: utf-8 -*-
"""
clean_hr_근태_2026_q1.py

`HR_근태_2026_1분기_원본수령분.xlsx` 정제 스크립트.
근거: HR_근태_2026_1분기_정제규칙.md (규칙 1~7)

시트 이름이 "YYYY년M월" 패턴이면 몇 개가 있든(1개월치든 12개월치든) 자동으로 인식합니다.
월 목록을 코드에 하드코딩하지 않았으므로, 나중에 시트가 추가/삭제돼도 스크립트 수정이 필요 없습니다.

사용법:
    python clean_hr_근태_2026_q1.py [원본.xlsx] [출력.csv]
    인자를 생략하면 아래 기본값(DEFAULT_SRC/DEFAULT_OUT)을 사용합니다.
    원본 파일과 같은 폴더에서 실행하면 별도 인자 없이 바로 동작합니다.
"""
import re
import sys
import pandas as pd
from openpyxl import load_workbook

DEFAULT_SRC = "HR_근태_2026_1분기_원본수령분.xlsx"
DEFAULT_OUT = "HR_근태_2026_1분기_정제.csv"

SHEET_NAME_PATTERN = re.compile(r"^(\d{4})년\s*(\d{1,2})월$")  # 예: "2026년1월" -> (2026, 1)

DEPT_WHITELIST = {"CS", "개발", "생산", "영업", "인사", "재무"}  # 규칙6 검증용

# 색상 -> 이상근태 플래그 매핑 (원본 생성 시 사용한 색상 코드)
COLOR_ABNORMAL = "00FFC7CE"  # 빨강
COLOR_NORMAL = "00C6EFCE"    # 초록


def parse_sheet_to_ym(sheet_name):
    """'2026년1월' -> '2026-01'. 패턴이 안 맞으면 None (해당 시트는 건너뜀).
    시트 개수·월 개수가 몇 개든(1개월치든 12개월치든) 하드코딩 없이 자동 인식한다."""
    m = SHEET_NAME_PATTERN.match(sheet_name.strip())
    if not m:
        return None
    year, month = m.groups()
    return f"{year}-{int(month):02d}"


def parse_overtime_text(v):
    """규칙3: '13.5시간' -> 13.5 (float). 패턴이 안 맞으면 None + 실패 로그용 원본값 반환."""
    if v is None:
        return None, None
    m = re.match(r"^\s*([\d.]+)\s*시간\s*$", str(v))
    if m:
        return float(m.group(1)), None
    return None, v  # 변환 실패 -> (None, 원본값)


def clean_sheet(ws, ym, parse_failures):
    rows, subtotal_rows = [], []
    current_dept = None

    for row in range(3, ws.max_row + 1):
        dept_cell = ws.cell(row=row, column=1).value
        emp_cell = ws.cell(row=row, column=2).value
        base_cell = ws.cell(row=row, column=3).value
        ot_cell = ws.cell(row=row, column=4).value
        note_cell = ws.cell(row=row, column=5)

        # 규칙1: 소계 행(그레인이 다름) 분리
        # -- 소계 행 자체의 "OO 소계" 텍스트가 current_dept를 덮어쓰기 전에 먼저 분리
        if dept_cell and "소계" in str(dept_cell):
            subtotal_rows.append(
                {"부서": current_dept, "년월": ym, "원본_소계행_텍스트": ot_cell}
            )
            continue

        # 규칙6: 병합 셀로 인한 부서 결측 -> forward-fill
        if dept_cell:
            current_dept = str(dept_cell)

        if emp_cell is None:
            continue  # 완전 빈 행 방지

        overtime, fail_raw = parse_overtime_text(ot_cell)  # 규칙3
        if fail_raw is not None:
            parse_failures.append({"사번": emp_cell, "년월": ym, "원본값": fail_raw})

        rgb = note_cell.fill.fgColor.rgb if note_cell.fill and note_cell.fill.fgColor else None
        if rgb == COLOR_ABNORMAL:
            abnormal_flag = True
        elif rgb == COLOR_NORMAL:
            abnormal_flag = False
        else:
            abnormal_flag = None  # 규칙5: 색상 자체가 없는 경우(원본에서는 발생하지 않아야 함)

        rows.append({
            "사번": emp_cell,
            "부서": current_dept,
            "년월": ym,
            "기본근무시간_가정값": base_cell,  # 규칙4: 관측치 아님을 컬럼명에 명시
            "초과근무시간": overtime,
            "원본이상근태플래그": abnormal_flag,  # 규칙5: 재판정 없이 원본 그대로 보존
        })

    return pd.DataFrame(rows), pd.DataFrame(subtotal_rows)


def main():
    src = sys.argv[1] if len(sys.argv) > 1 else DEFAULT_SRC
    out = sys.argv[2] if len(sys.argv) > 2 else DEFAULT_OUT

    wb = load_workbook(src)
    all_rows, all_subtotals, parse_failures = [], [], []
    detected_yms = []

    for sheet_name in wb.sheetnames:
        ym = parse_sheet_to_ym(sheet_name)
        if ym is None:
            print(f"⚠️ 시트명이 'YYYY년M월' 패턴이 아니라 건너뜀: {sheet_name!r}")
            continue
        detected_yms.append(ym)
        ws = wb[sheet_name]
        df, subtotal_df = clean_sheet(ws, ym, parse_failures)
        all_rows.append(df)
        all_subtotals.append(subtotal_df)

    if not all_rows:
        print("❌ 'YYYY년M월' 패턴에 맞는 시트를 하나도 찾지 못했습니다.")
        return None, None

    cleaned = pd.concat(all_rows, ignore_index=True)
    subtotals = pd.concat(all_subtotals, ignore_index=True)

    print("=" * 60)
    print(f"자동 인식된 월 시트: {detected_yms} (총 {len(detected_yms)}개 — 하드코딩 아님, 시트명에서 자동 파싱)")
    print("\n규칙 1 — 그레인 분리")
    print(f"  상세 행(사번×월): {len(cleaned)}행 / 소계 행(부서×월): {len(subtotals)}행")

    print("\n규칙 2 — 년월 컬럼 검증")
    ym_ok = cleaned["년월"].isin(detected_yms).all() and cleaned["년월"].notna().all()
    print(f"  년월 값 전부 유효({sorted(set(detected_yms))}) 및 결측 없음: {'✅' if ym_ok else '❌'}")

    print("\n규칙 3 — 텍스트→숫자 변환 검증")
    if parse_failures:
        print(f"  ⚠️ 변환 실패 {len(parse_failures)}건 발견:")
        for f in parse_failures[:10]:
            print(f"    {f}")
    else:
        print("  ✅ 변환 실패 0건 (전체 정상 변환)")

    print("\n규칙 4 — 기본근무시간_가정값 컬럼 표시 확인")
    print(f"  컬럼 존재: {'기본근무시간_가정값' in cleaned.columns} (지표 계산에는 미사용 — 별도 컬럼으로 보존만 함)")

    print("\n규칙 5 — 이상근태 플래그 분포")
    print(cleaned.groupby("년월")["원본이상근태플래그"].value_counts().to_string())

    print("\n규칙 6 — 부서 화이트리스트 검증")
    bad_dept = set(cleaned["부서"].unique()) - DEPT_WHITELIST
    missing_dept = (cleaned["부서"].isna()).sum()
    print(f"  화이트리스트 외 부서: {bad_dept if bad_dept else '없음 ✅'} / 부서 결측: {missing_dept}건")

    print("\n규칙 7 — 소계 행 대조 검증")
    recomputed = (
        cleaned.groupby(["부서", "년월"])["초과근무시간"].mean().round(1).reset_index()
        .rename(columns={"초과근무시간": "재계산_평균"})
    )
    subtotals["원본_평균"] = subtotals["원본_소계행_텍스트"].str.extract(r"평균\s*([\d.]+)시간").astype(float)
    check = subtotals.merge(recomputed, on=["부서", "년월"])
    check["일치여부"] = (check["원본_평균"] - check["재계산_평균"]).abs() < 0.05
    print(check[["부서", "년월", "원본_평균", "재계산_평균", "일치여부"]].to_string(index=False))
    mismatch = check[~check["일치여부"]]
    if len(mismatch) > 0:
        print(f"\n  ⚠️ 불일치 {len(mismatch)}건 발견 — 정제 로직 재확인 필요")
    else:
        print(f"\n  ✅ 전체 {len(check)}건 일치 (오차 0.05시간 미만)")

    cleaned.to_csv(out, index=False, encoding="utf-8-sig")
    print("=" * 60)
    print(f"저장 완료: {out} ({len(cleaned)}행)")
    return cleaned, check


if __name__ == "__main__":
    main()
